import numpy as np
import re
import itertools
from collections import Counter
from openpyxl import Workbook,load_workbook
from datetime import datetime, timedelta

def clean_str(string):
    """
    Tokenization/string cleaning for all datasets except for SST.
    Original taken from https://github.com/yoonkim/CNN_sentence/blob/master/process_data.py
    """
    string = re.sub(r"[^A-Za-z0-9(),!?\'\`]", " ", string)
    string = re.sub(r"'", " ' ", string)
    string = re.sub(r"\'s", " \'s", string)
    string = re.sub(r"\'ve", " \'ve", string)
    string = re.sub(r"n\'t", " n\'t", string)
    string = re.sub(r"\'re", " \'re", string)
    string = re.sub(r"\'d", " \'d", string)
    string = re.sub(r"\'ll", " \'ll", string)
    string = re.sub(r",", " , ", string)
    string = re.sub(r"!", " ! ", string)
    string = re.sub(r"\(", " ( ", string)
    string = re.sub(r"\)", " ) ", string)
    string = re.sub(r"\?", " ? ", string)
    string = re.sub(r"\s{2,}", " ", string)
    return string.strip().lower()


def news_weight_dist(period):
    from scipy.stats import skewnorm
    from scipy.stats import gamma
    import matplotlib.pyplot as plt
    fig, ax = plt.subplots(1, 1)
    a=1.5
    x = np.linspace(0,1,period)
    dist = gamma.pdf(x, a,0,0.13)
    # print (dist/period)
    # print (np.sum(dist)/period)
    # ax.plot(x, dist,'r-', lw=3, alpha=1, label='skewnorm pdf')
    # plt.show()
    return (dist/period)

def work_days(ws_o):
    wrk_days = []
    for date in range(ws_o.max_row-1):
        date_o = datetime.strptime(str(ws_o['A'+str(date+2)+''].value), '%Y-%m-%d %H:%M:%S')
        if date_o.date() not in wrk_days:
            wrk_days.append(date_o.date())
    return wrk_days

def get_sentiment_val(ws_o,date_i,time_mins,ret_val):
    for j in range(ws_o.max_row-1):
        date_o = datetime.strptime(str(ws_o['A'+str(j+2)+''].value), '%Y-%m-%d %H:%M:%S')
        if format(date_o,'%Y-%m-%d %H:%M') == format(date_i,'%Y-%m-%d %H:%M'):
            ret_val = [float(ws_o['C'+str(j+2+x)+''].value) if ws_o['C'+str(j+2+x)+''].value else 0 for x in range(time_mins)]
            break
    return ret_val

def sentiment_score(ws_o,date_i,date_next,time_mins,wrk_days,last,i):
    ret_val,get_data = [],True
    # All single headlines within trading day and time i.e. Monday - Friday 9:30 to 4:29
    if (date_i.date() in wrk_days) and (date_i.time() >= datetime.strptime('9:30', '%H:%M').time()) and (date_i.time() <=datetime.strptime('16:29', '%H:%M').time()):
        ret_val = get_sentiment_val(ws_o,date_i,time_mins,ret_val)
        get_data = False
    if get_data:
        # For the days when next day is a working day
        if (date_next.date() in wrk_days) and (date_next.time() >= datetime.strptime('9:30', '%H:%M').time()) and (date_i.time() <=datetime.strptime('16:29', '%H:%M').time()):
            ret_val = get_sentiment_val(ws_o, datetime.strptime(str(date_next.date()) + ' 9:30', '%Y-%m-%d %H:%M'),time_mins,ret_val)
        if i == last:
            ret_val = get_sentiment_val(ws_o, datetime.strptime(str(date_next.date()) + ' 9:30', '%Y-%m-%d %H:%M'),time_mins,ret_val)
    return ret_val

def load_data(data_file, output_file, article_text = True):
    wb_i,wb_o = load_workbook(data_file),load_workbook(output_file,data_only = True)
    ws_i,ws_o = wb_i['AAPL NEWS'],wb_o['Sheet1']
    time_mins = 240
    weight_dist = news_weight_dist(time_mins)
    wrk_days = work_days(ws_o)
    print(wrk_days)
    input_x,text_data= {},''
    for i in range(ws_i.max_row-2):
        last = ws_i.max_row-3
        try :
            date_i = datetime.strptime(str(ws_i['A'+str(i+2)+''].value),'%Y-%m-%d %H:%M:%S.%f')
        except:
            date_i = datetime.strptime(str(ws_i['A'+str(i+2)+''].value),'%Y-%m-%d %H:%M:%S')
        try :
            date_next = datetime.strptime(str(ws_i['A'+str(i+3)+''].value),'%Y-%m-%d %H:%M:%S.%f')
        except:
            date_next = datetime.strptime(str(ws_i['A'+str(i+3)+''].value),'%Y-%m-%d %H:%M:%S')
        headline = clean_str(re.sub(r'[^\x00-\x7F]+',' ', str(ws_i['D'+str(i+2)+''].value)))
        # print ('\n\n=> News : ',headline)
        text = clean_str(re.sub(r'[^\x00-\x7F]+',' ', str(ws_i['E'+str(i+2)+''].value)))
        if article_text:
        	text_data +=' '+ headline + ' ' + text
        else :
        	text_data +=' '+ headline
        ret_val = sentiment_score(ws_o,date_i,date_next,time_mins,wrk_days,last,i)
        if ret_val:
            ret_val = np.array(ret_val,dtype=float)
            ret_val = [i/sum(abs(ret_val)) for i in ret_val]
            ret_val = np.array(ret_val,dtype=float)
            senti_val = np.sum(np.multiply(weight_dist,ret_val))
            input_x[text_data] = senti_val*1000
            # print ('Sentiment : ',senti_val)
            text_data= ''
        else :
            pass
    return (input_x)

def get_data(data_file, output_file, article_text = True):
    x_text = load_data(data_file, output_file, article_text )
    input_x,y = [],[]
    for key,value in x_text.items():
        input_x.append(key)
        y.append([1,0] if value >0 else [0,1])
    return (input_x,y)

def load_test(excel_file):
    '''
    Loads the test excel file with 100 headlines marked with positive and negative sentiment
    Process the excel file and returns the headline and sentiment
    '''
    wb = load_workbook('test.xlsx')
    worksheet = wb['Sheet1']
    x_text,y = [],[]
    #print (worksheet['C1'].value)
    for i in range(100):
        senti_val = str(worksheet['B'+str(i+1)+''].value)
        if (senti_val == 'n' or senti_val=='p'):
            headline = str(worksheet['A'+str(i+1)+''].value)
            line = re.sub(r'[^\x00-\x7F]+',' ', headline)
            y.append([0,1] if senti_val == 'p' else [1,0])
            x_text.append(line)
    x_text = [clean_str(sent) for sent in x_text]
    return [x_text,np.array(y)]

def process_vocab(dim,word_vec):
    '''
    Processes the word vectors from glove.
    Returns vocabulary and the corresponding word-vectors.
    '''
    vocab,word2vec = {},[[0.0]*dim]
    for idx,(word,vec) in enumerate(word_vec.items()):
            vocab[word] = idx+1
            word2vec.append(list(map(float, vec)))
    print("==> vocab processing is done")
    return vocab, word2vec

def load_glove(dim):
    '''
    Loads the GloVe word-vectors with mentioned dimension.
    Return the word vectors in  dictionary format.
    '''
    word_vec = {}
    print("==> loading glove")
    with open(("glove.6B." + str(dim) + "d.txt" ), encoding = 'utf8') as f:  # "vec"
        for line in f:
            l = line.split()
            for idx,word in enumerate(l):
                if idx !=0:
                    try :
                        float(word)
                        break
                    except :
                        pass
            key = ' '.join(l[:idx])
            # print (key.encode("utf8"))
            word_vec[key] = list(map(float, l[idx:]))
            
    print("==> glove is loaded")
    return word_vec

def create_vector(word,dim,word2vec):
    '''
    Creates a word vector if the word is not present in the vocab.
    Returns the updated word vectors.
    '''
    vector = np.random.uniform(0.0,1.0,(dim,))
    word2vec[word] = vector
    with open("glove.6B." + str(dim) + "d.txt" ,'a') as gw:
    	gw.write('\n'+str(word))
    	for vec in vector.tolist():
    		gw.write(' '+str(vec))
    return word2vec

def batch_iter(data, batch_size, num_epochs, shuffle=True):
    """
    Generates a batch iterator for a dataset.
    """
    data = np.array(data)
    data_size = len(data)
    num_batches_per_epoch = int((len(data)-1)/batch_size) + 1
    for epoch in range(num_epochs):
        # Shuffle the data at each epoch
        if shuffle:
            shuffle_indices = np.random.permutation(np.arange(data_size))
            shuffled_data = data[shuffle_indices]
        else:
            shuffled_data = data
        for batch_num in range(num_batches_per_epoch):
            start_index = batch_num * batch_size
            end_index = min((batch_num + 1) * batch_size, data_size)
            yield shuffled_data[start_index:end_index]