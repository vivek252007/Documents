import data_helpers as dh
from openpyxl import Workbook,load_workbook
from datetime import datetime
import numpy as np

def change_excel(data_file):
    wb_i = load_workbook(data_file)
    ws_i = wb_i['AAPL NEWS']
    for i in range(ws_i.max_row -1):
        date_obj = datetime.strptime(str(ws_i['A'+str(i+2)+''].value)[:10]+' '+str(ws_i['B'+str(i+2)+''].value),'%Y-%m-%d %H:%M:%S')
        print (date_obj)
        ws_i['A'+str(i+2)+''] = date_obj
    wb_i.save(data_file)

# print (dh.get_data('AAPL.xlsx','AAPL_min.xlsx') )

def news_weight_dist(period):
    from scipy.stats import skewnorm
    from scipy.stats import gamma
    import matplotlib.pyplot as plt
    fig, ax = plt.subplots(1, 1)
    a=1.5
    x = np.linspace(0,1,period)
    dist = gamma.pdf(x, a,0,0.13)
    print (dist/period)
    print (np.sum(dist)/period)
    ax.plot(x, dist,'r-', lw=3, alpha=1, label='skewnorm pdf')
    plt.show()
    return (dist/period)

import plotly.plotly as py
from plotly.graph_objs import *

trace0 = Scatter(
    x=[1, 2, 3, 4],
    y=[10, 15, 13, 17]
)
trace1 = Scatter(
    x=[1, 2, 3, 4],
    y=[16, 5, 11, 9]
)
data = Data([trace0, trace1])

py.plot(data, filename = 'basic-line')

query = """SELECT t1.storyidentifier, t1.tickers, t1.storytime, t1.storyheadline t2.bodytext FROM t_bb_news_story_analytics_v2 t1 
INNER JOIN t_bb_story_text_indexed t2 ON t2.storyidentifier = t1.storyidentifier limit 500"""

df_raw_copper = psql.read_sql(query, conn)
df_raw_copper.to_csv("database.csv", sep='|', encoding='utf-8')

DatabaseError: Execution failed on sql 'SELECT t1.storyidentifier, t1.tickers, t1.storytime, t1.storyheadline t2.bodytext FROM t_bb_news_story_analytics_v2 t1 
INNER JOIN t_bb_story_text_indexed t2 ON t2.storyidentifier = t1.storyidentifier limit 500': syntax error at or near "."
LINE 1: ...ier, t1.tickers, t1.storytime, t1.storyheadline t2.bodytext ...
